import json
import os
from io import BytesIO
from functools import wraps
from flask import Flask, render_template, request, redirect, url_for, flash, session, send_file, g
from pathlib import Path
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from markupsafe import Markup, escape
from werkzeug.security import check_password_hash, generate_password_hash

try:
    from .db import column_names, get_db_connection, sql_type, table_exists
except ImportError:
    from db import column_names, get_db_connection, sql_type, table_exists

BASE_DIR = Path(__file__).resolve().parent
DB_PATH = BASE_DIR / "inventory.db"

app = Flask(__name__)
app.secret_key = os.getenv("SECRET_KEY", "dev-inventory-secret-key")

OPERATION_LABELS = {
    "inbound": "入庫",
    "outbound": "出庫",
    "adjustment": "棚卸",
    "modification": "修正",
}

PRODUCT_COLUMNS = {
    "source_sheet": "TEXT",
    "source_row": "INTEGER",
    "big_category": "TEXT",
    "maker_or_product": "TEXT",
    "overview": "TEXT",
    "supplier": "TEXT",
    "amount": "TEXT",
    "stock_status": "TEXT",
    "display_flag": "TEXT",
    "available_stock": "INTEGER NOT NULL DEFAULT 0",
    "total_stock": "INTEGER NOT NULL DEFAULT 0",
    "staff_stock_json": "TEXT",
    "imported_at": "TEXT",
}

STOCK_LOG_COLUMNS = {
    "metadata_json": "TEXT",
}

USER_COLUMNS = {
    "username": "TEXT NOT NULL UNIQUE",
    "password_hash": "TEXT NOT NULL",
    "role": "TEXT NOT NULL DEFAULT 'user'",
    "is_active": "INTEGER NOT NULL DEFAULT 1",
    "admin_requested": "INTEGER NOT NULL DEFAULT 0",
    "created_at": "TEXT NOT NULL",
}


def ensure_product_columns(conn):
    existing_columns = column_names("products")
    for name, definition in PRODUCT_COLUMNS.items():
        if name not in existing_columns:
            conn.execute(f"ALTER TABLE products ADD COLUMN {name} {definition}")
    conn.commit()


def ensure_stock_log_columns(conn):
    existing_columns = column_names("stock_logs")
    for name, definition in STOCK_LOG_COLUMNS.items():
        if name not in existing_columns:
            conn.execute(f"ALTER TABLE stock_logs ADD COLUMN {name} {definition}")
    conn.commit()


def ensure_alert_tables(conn):
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS low_stock_alerts (
            id {pk},
            product_id INTEGER NOT NULL UNIQUE,
            threshold INTEGER NOT NULL DEFAULT 0,
            FOREIGN KEY(product_id) REFERENCES products(id) ON DELETE CASCADE
        )
        """.format(pk=sql_type("pk"))
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS zero_stock_alerts (
            id {pk},
            product_id INTEGER NOT NULL UNIQUE,
            FOREIGN KEY(product_id) REFERENCES products(id) ON DELETE CASCADE
        )
        """.format(pk=sql_type("pk"))
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS stagnant_stock_alerts (
            id {pk},
            product_id INTEGER NOT NULL UNIQUE,
            FOREIGN KEY(product_id) REFERENCES products(id) ON DELETE CASCADE
        )
        """.format(pk=sql_type("pk"))
    )
    conn.commit()


def ensure_user_table(conn):
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS users (
            id {pk},
            username TEXT NOT NULL UNIQUE,
            password_hash TEXT NOT NULL,
            role TEXT NOT NULL DEFAULT 'user',
            is_active INTEGER NOT NULL DEFAULT 1,
            admin_requested INTEGER NOT NULL DEFAULT 0,
            created_at TEXT NOT NULL
        )
        """.format(pk=sql_type("pk"))
    )
    conn.commit()
    existing_columns = column_names("users")
    for name, definition in USER_COLUMNS.items():
        if name not in existing_columns:
            conn.execute(f"ALTER TABLE users ADD COLUMN {name} {definition}")
    conn.commit()


def ensure_initial_admin(conn):
    username = os.getenv("ADMIN_USERNAME")
    password = os.getenv("ADMIN_PASSWORD")
    if not username or not password:
        return
    existing = conn.execute("SELECT id FROM users WHERE username = ?", (username,)).fetchone()
    if existing:
        return
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn.execute(
        "INSERT INTO users (username, password_hash, role, is_active, admin_requested, created_at) VALUES (?, ?, ?, ?, ?, ?)",
        (username, generate_password_hash(password), "admin", 1, 0, now),
    )
    conn.commit()


def ensure_database():
    conn = get_db_connection()
    try:
        if table_exists("products"):
            ensure_product_columns(conn)
            ensure_stock_log_columns(conn)
            ensure_alert_tables(conn)
        else:
            conn.close()
            conn = None
            try:
                from .init_db import create_db
            except ImportError:
                from init_db import create_db
            create_db()
            conn = get_db_connection()
        ensure_user_table(conn)
        ensure_initial_admin(conn)
    finally:
        if conn:
            conn.close()


ensure_database()


def format_staff_stock_display(staff_stock_json):
    """
    Filter to display staff sales without zero values.
    Input: JSON string like '{"A05": 1, "A06": 0, "H12": 2}'
    Output: each staff stock on its own line, or "-" if empty or all zeros
    """
    if not staff_stock_json:
        return "-"
    try:
        data = json.loads(staff_stock_json)
        non_zero = {k: v for k, v in data.items() if v != 0}
        if not non_zero:
            return "-"
        return Markup("<br>".join([f"{escape(k)}: {escape(v)}" for k, v in non_zero.items()]))
    except Exception:
        return "-"


app.jinja_env.filters['format_staff_stock'] = format_staff_stock_display


@app.before_request
def load_current_user():
    g.user = None
    user_id = session.get("user_id")
    if not user_id:
        return
    conn = get_db_connection()
    try:
        g.user = conn.execute(
            "SELECT id, username, role FROM users WHERE id = ?",
            (user_id,),
        ).fetchone()
    finally:
        conn.close()


@app.context_processor
def inject_auth_context():
    return {
        "current_user": g.get("user"),
        "is_admin": bool(g.get("user") and g.user["role"] == "admin"),
    }


def login_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if g.get("user") is None:
            return redirect(url_for("login", next=request.full_path))
        return view(*args, **kwargs)
    return wrapped


def admin_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if g.get("user") is None:
            return redirect(url_for("login", next=request.full_path))
        if g.user["role"] != "admin":
            flash("管理者権限が必要です。", "danger")
            return redirect(url_for("dashboard"))
        return view(*args, **kwargs)
    return wrapped


@app.route("/login", methods=["GET", "POST"])
def login():
    if g.get("user"):
        return redirect(url_for("dashboard"))

    error_message = None
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        conn = get_db_connection()
        try:
            user = conn.execute("SELECT * FROM users WHERE username = ?", (username,)).fetchone()
        finally:
            conn.close()
        if user and not user["is_active"]:
            error_message = "このアカウントは無効です。管理者に確認してください。"
        elif user and check_password_hash(user["password_hash"], password):
            session.clear()
            session["user_id"] = user["id"]
            session["role"] = user["role"]
            return redirect(request.args.get("next") or url_for("dashboard"))
        else:
            error_message = "ユーザー名またはパスワードが正しくありません。"

    return render_template("login.html", error_message=error_message)


@app.route("/register", methods=["GET", "POST"])
def register():
    if g.get("user"):
        return redirect(url_for("dashboard"))

    error_message = None
    form_data = {}
    if request.method == "POST":
        form_data = request.form.to_dict()
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        password_confirm = request.form.get("password_confirm", "")
        admin_requested = 1 if request.form.get("admin_requested") == "1" else 0

        if not username:
            error_message = "ユーザー名を入力してください。"
        elif len(password) < 8:
            error_message = "パスワードは8文字以上で入力してください。"
        elif password != password_confirm:
            error_message = "確認用パスワードが一致しません。"
        else:
            conn = get_db_connection()
            try:
                existing = conn.execute("SELECT id FROM users WHERE username = ?", (username,)).fetchone()
                if existing:
                    error_message = "同じユーザー名がすでに使われています。"
                else:
                    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    conn.execute(
                        "INSERT INTO users (username, password_hash, role, is_active, admin_requested, created_at) VALUES (?, ?, ?, ?, ?, ?)",
                        (username, generate_password_hash(password), "user", 1, admin_requested, now),
                    )
                    conn.commit()
                    flash("アカウントを作成しました。ログインしてください。", "success")
                    return redirect(url_for("login"))
            finally:
                conn.close()

    return render_template("register.html", error_message=error_message, form_data=form_data)


@app.route("/logout")
def logout():
    session.clear()
    flash("ログアウトしました。", "success")
    return redirect(url_for("login"))


@app.route("/users", methods=["GET", "POST"])
@admin_required
def users():
    conn = get_db_connection()
    if request.method == "POST":
        user_id = request.form.get("user_id", type=int)
        action = request.form.get("action")
        target = conn.execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()
        if target is None:
            flash("ユーザーが見つかりません。", "danger")
        elif target["id"] == g.user["id"] and action in ["deactivate", "make_user"]:
            flash("自分自身の管理者権限や有効状態は変更できません。", "danger")
        elif action == "make_admin":
            conn.execute(
                "UPDATE users SET role = ?, admin_requested = ? WHERE id = ?",
                ("admin", 0, user_id),
            )
            conn.commit()
            flash("管理者権限を付与しました。", "success")
        elif action == "make_user":
            conn.execute(
                "UPDATE users SET role = ?, admin_requested = ? WHERE id = ?",
                ("user", 0, user_id),
            )
            conn.commit()
            flash("普通ユーザーに変更しました。", "success")
        elif action == "activate":
            conn.execute("UPDATE users SET is_active = ? WHERE id = ?", (1, user_id))
            conn.commit()
            flash("ユーザーを有効にしました。", "success")
        elif action == "deactivate":
            conn.execute("UPDATE users SET is_active = ? WHERE id = ?", (0, user_id))
            conn.commit()
            flash("ユーザーを無効にしました。", "success")
        else:
            flash("操作が無効です。", "danger")
        conn.close()
        return redirect(url_for("users"))

    rows = conn.execute("SELECT id, username, role, is_active, admin_requested, created_at FROM users ORDER BY created_at DESC, username ASC").fetchall()
    conn.close()
    return render_template("users.html", users=rows)


def parse_search_args():
    return {
        "q": request.args.get("q", "", type=str).strip(),
        "sku": request.args.get("sku", "", type=str).strip(),
        "big_category": request.args.get("big_category", "", type=str).strip(),
        "location": request.args.get("location", "", type=str).strip(),
    }


def build_product_query(params):
    sql = "SELECT * FROM products WHERE 1=1"
    values = []

    if params["q"]:
        sql += " AND (name LIKE ? OR sku LIKE ? OR big_category LIKE ? OR maker_or_product LIKE ? OR overview LIKE ? OR notes LIKE ? OR source_sheet LIKE ? OR category LIKE ? OR location LIKE ? OR staff_stock_json LIKE ? )"
        term = f"%{params['q']}%"
        values.extend([term] * 10)

    if params["sku"]:
        sql += " AND sku LIKE ?"
        values.append(f"%{params['sku']}%")

    if params["big_category"]:
        sql += " AND big_category = ?"
        values.append(params["big_category"])

    if params["location"]:
        sql += " AND location LIKE ?"
        values.append(f"%{params['location']}%")

    sql += " ORDER BY name ASC"
    return sql, values


def parse_int(value):
    if value is None:
        return 0
    if isinstance(value, str):
        value = value.strip().replace(",", "")
    try:
        return int(value)
    except Exception:
        try:
            return int(float(value))
        except Exception:
            return 0


def parse_staff_stock_form():
    staff_stock = {}
    names = request.form.getlist("staff_name")
    quantities = request.form.getlist("staff_quantity")
    for name, quantity in zip(names, quantities):
        name = name.strip()
        stock = parse_int(quantity)
        if name and stock != 0:
            staff_stock[name] = stock
    return staff_stock


def staff_stock_form_rows(staff_stock_json=None, min_rows=3):
    rows = []
    if staff_stock_json:
        try:
            rows = [
                {"name": name, "quantity": quantity}
                for name, quantity in json.loads(staff_stock_json).items()
            ]
        except Exception:
            rows = []
    while len(rows) < min_rows:
        rows.append({"name": "", "quantity": ""})
    return rows


def adjust_staff_sales_json(staff_stock_json, staff_name, delta):
    if not staff_name or delta == 0:
        return staff_stock_json or "{}"
    try:
        data = json.loads(staff_stock_json or "{}")
    except Exception:
        data = {}
    current = parse_int(data.get(staff_name, 0))
    data[staff_name] = max(0, current + delta)
    return json.dumps(data, ensure_ascii=False)


PRODUCT_SNAPSHOT_FIELDS = [
    "source_sheet",
    "big_category",
    "maker_or_product",
    "overview",
    "supplier",
    "amount",
    "sku",
    "available_stock",
    "current_stock",
    "total_stock",
    "staff_stock_json",
    "notes",
    "name",
    "category",
    "location",
]


def product_snapshot(product):
    return {field: product[field] for field in PRODUCT_SNAPSHOT_FIELDS if field in product.keys()}


def restore_product_snapshot(conn, product_id, snapshot):
    fields = [field for field in PRODUCT_SNAPSHOT_FIELDS if field in snapshot]
    if not fields:
        return False
    assignments = ", ".join([f"{field} = ?" for field in fields])
    values = [snapshot[field] for field in fields]
    values.extend([datetime.now().strftime("%Y-%m-%d %H:%M:%S"), product_id])
    conn.execute(
        f"UPDATE products SET {assignments}, updated_at = ? WHERE id = ?",
        values,
    )
    return True


def get_inventory_form_options(conn):
    categories = [
        row[0] for row in conn.execute(
            "SELECT DISTINCT big_category FROM products WHERE big_category IS NOT NULL AND big_category != '' ORDER BY big_category"
        ).fetchall()
    ]
    makers = [
        row[0] for row in conn.execute(
            "SELECT DISTINCT maker_or_product FROM products WHERE maker_or_product IS NOT NULL AND maker_or_product != '' ORDER BY maker_or_product"
        ).fetchall()
    ]
    source_sheets = [
        row[0] for row in conn.execute(
            "SELECT DISTINCT source_sheet FROM products WHERE source_sheet IS NOT NULL AND source_sheet != '' ORDER BY source_sheet"
        ).fetchall()
    ]
    staff_names = set()
    rows = conn.execute(
        "SELECT staff_stock_json FROM products WHERE staff_stock_json IS NOT NULL AND staff_stock_json != ''"
    ).fetchall()
    for row in rows:
        try:
            staff_names.update(json.loads(row[0]).keys())
        except Exception:
            continue
    return {
        "categories": categories,
        "makers": makers,
        "source_sheets": source_sheets,
        "staff_names": sorted(staff_names),
        "quantities": list(range(0, 21)),
    }


def compute_stock_delta(operation_type, quantity):
    if operation_type == "inbound":
        return abs(quantity)
    if operation_type == "outbound":
        return -abs(quantity)
    return quantity


def get_cell(sheet, row, col):
    value = sheet.cell(row=row, column=col).value
    return value if value is not None else ""


def parse_staff_stock(sheet, row):
    values = {}
    for col in range(13, 27):
        header = sheet.cell(row=2, column=col).value
        if header is None:
            continue
        header = str(header).strip()
        if not header:
            continue
        cell_value = sheet.cell(row=row, column=col).value
        stock = parse_int(cell_value)
        values[header] = stock
    return values


NORMAL_SHEETS = {
    "エアコン在庫": "エアコン",
    "給湯器在庫": "給湯器",
    "トイレ在庫": "トイレ",
    "その他設備在庫": None,
    "その他家電在庫": None,
}


def parse_inventory_row(sheet, sheet_name, row):
    if sheet_name in NORMAL_SHEETS:
        reorder_point = parse_int(get_cell(sheet, row, 2))
        stock_status = str(get_cell(sheet, row, 3)).strip()
        maker_or_product = str(get_cell(sheet, row, 4)).strip()
        overview = str(get_cell(sheet, row, 5)).strip()
        sku = str(get_cell(sheet, row, 6)).strip()
        display_flag = str(get_cell(sheet, row, 7)).strip()
        available_stock = parse_int(get_cell(sheet, row, 8))
        supplier = str(get_cell(sheet, row, 9)).strip()
        amount = str(get_cell(sheet, row, 10)).strip()
        total_stock = parse_int(get_cell(sheet, row, 12))
        notes = str(get_cell(sheet, row, 27)).strip()
        staff_stock = parse_staff_stock(sheet, row)

        if not sku and not maker_or_product and not overview:
            return None

        if NORMAL_SHEETS[sheet_name] is None:
            big_category = maker_or_product
        else:
            big_category = NORMAL_SHEETS[sheet_name]

        name = maker_or_product or sku or overview or "不明"
        return {
            "source_sheet": sheet_name,
            "source_row": row,
            "big_category": big_category,
            "maker_or_product": maker_or_product,
            "overview": overview,
            "supplier": supplier,
            "amount": amount,
            "sku": sku,
            "stock_status": stock_status,
            "display_flag": display_flag,
            "available_stock": available_stock,
            "total_stock": total_stock,
            "reorder_point": reorder_point,
            "staff_stock_json": json.dumps(staff_stock, ensure_ascii=False),
            "notes": notes,
            "imported_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "current_stock": available_stock,
            "name": name,
            "category": big_category,
            "location": "",
        }
    if sheet_name == "魚倉庫在庫":
        reorder_point = parse_int(get_cell(sheet, row, 2))
        stock_status = str(get_cell(sheet, row, 3)).strip()
        available_stock = parse_int(get_cell(sheet, row, 4))
        maker_or_product = str(get_cell(sheet, row, 8)).strip()
        overview = str(get_cell(sheet, row, 7)).strip()
        sku = str(get_cell(sheet, row, 9)).strip()
        supplier = str(get_cell(sheet, row, 10)).strip()
        amount = str(get_cell(sheet, row, 11)).strip()
        notes = str(get_cell(sheet, row, 20)).strip()
        big_category = str(get_cell(sheet, row, 6)).strip()

        if not sku and not overview:
            return None

        name = maker_or_product or sku or overview or "不明"
        return {
            "source_sheet": sheet_name,
            "source_row": row,
            "big_category": big_category,
            "maker_or_product": maker_or_product,
            "overview": overview,
            "supplier": supplier,
            "amount": amount,
            "sku": sku,
            "stock_status": stock_status,
            "display_flag": "",
            "available_stock": available_stock,
            "total_stock": 0,
            "reorder_point": reorder_point,
            "staff_stock_json": json.dumps({}, ensure_ascii=False),
            "notes": notes,
            "imported_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "current_stock": available_stock,
            "name": name,
            "category": big_category,
            "location": "",
        }
    return None


def import_excel_file(file_stream):
    workbook = load_workbook(filename=file_stream, data_only=True)
    sheet_names = [name for name in workbook.sheetnames if workbook[name].sheet_state == "visible"]
    imported = 0
    skipped = 0
    errors = 0
    error_details = []
    deleted_products = 0
    deleted_logs = 0
    
    conn = get_db_connection()
    
    try:
        # Delete old data before importing
        cursor = conn.cursor()
        cursor.execute("DELETE FROM stock_logs")
        deleted_logs = cursor.rowcount
        cursor.execute("DELETE FROM products")
        deleted_products = cursor.rowcount
        conn.commit()
        
        # Now import new data
        for sheet_name in sheet_names:
            if sheet_name not in list(NORMAL_SHEETS.keys()) + ["魚倉庫在庫"]:
                continue
            sheet = workbook[sheet_name]
            for row in range(4, sheet.max_row + 1):
                try:
                    record = parse_inventory_row(sheet, sheet_name, row)
                    if record is None:
                        skipped += 1
                        continue
                    
                    conn.execute(
                        "INSERT INTO products (source_sheet, source_row, big_category, maker_or_product, overview, supplier, amount, sku, stock_status, display_flag, available_stock, total_stock, reorder_point, staff_stock_json, notes, imported_at, current_stock, name, category, location, created_at, updated_at) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                        (
                            record["source_sheet"],
                            record["source_row"],
                            record["big_category"],
                            record["maker_or_product"],
                            record["overview"],
                            record["supplier"],
                            record["amount"],
                            record["sku"],
                            record["stock_status"],
                            record["display_flag"],
                            record["available_stock"],
                            record["total_stock"],
                            record["reorder_point"],
                            record["staff_stock_json"],
                            record["notes"],
                            record["imported_at"],
                            record["current_stock"],
                            record["name"],
                            record["category"],
                            record["location"],
                            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        ),
                    )
                    imported += 1
                except Exception as exc:
                    errors += 1
                    error_details.append(f"{sheet_name} 行 {row}: {exc}")
        
        conn.commit()
    except Exception as exc:
        conn.rollback()
        raise exc
    finally:
        conn.close()
    
    return {
        "imported": imported,
        "skipped": skipped,
        "errors": errors,
        "error_details": error_details,
        "deleted_products": deleted_products,
        "deleted_logs": deleted_logs,
    }


def month_range(target_date=None):
    target_date = target_date or datetime.now()
    start = target_date.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    if start.month == 12:
        end = start.replace(year=start.year + 1, month=1)
    else:
        end = start.replace(month=start.month + 1)
    return start, end


def stock_log_delta(operation_type, quantity):
    quantity = parse_int(quantity)
    if operation_type == "inbound":
        return abs(quantity)
    if operation_type == "outbound":
        return -abs(quantity)
    return quantity


def style_report_sheet(sheet):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.freeze_panes = "A2"
    for column in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column)
        sheet.column_dimensions[column[0].column_letter].width = min(max(max_length + 2, 10), 32)


def export_monthly_changes_file(target_date=None):
    start, end = month_range(target_date)
    conn = get_db_connection()
    logs = conn.execute(
        """
        SELECT
            stock_logs.*,
            products.sku,
            products.big_category,
            products.maker_or_product,
            products.overview,
            products.current_stock
        FROM stock_logs
        JOIN products ON stock_logs.product_id = products.id
        WHERE stock_logs.created_at >= ? AND stock_logs.created_at < ?
        ORDER BY stock_logs.created_at ASC, stock_logs.id ASC
        """,
        (start.strftime("%Y-%m-%d %H:%M:%S"), end.strftime("%Y-%m-%d %H:%M:%S")),
    ).fetchall()
    conn.close()

    workbook = Workbook()
    detail_sheet = workbook.active
    detail_sheet.title = "明細"
    detail_sheet.append([
        "日時",
        "操作",
        "品番",
        "大分類",
        "メーカー/商品",
        "概要",
        "数量",
        "増減",
        "担当者",
        "メモ",
        "現在販売可能台数",
    ])

    summary = {}
    for log in logs:
        delta = stock_log_delta(log["operation_type"], log["quantity"])
        label = OPERATION_LABELS.get(log["operation_type"], log["operation_type"])
        detail_sheet.append([
            log["created_at"],
            label,
            log["sku"],
            log["big_category"],
            log["maker_or_product"],
            log["overview"],
            abs(parse_int(log["quantity"])) if log["operation_type"] in ["inbound", "outbound"] else log["quantity"],
            delta,
            log["staff_name"],
            log["note"],
            log["current_stock"],
        ])

        key = log["product_id"]
        if key not in summary:
            summary[key] = {
                "sku": log["sku"],
                "big_category": log["big_category"],
                "maker_or_product": log["maker_or_product"],
                "overview": log["overview"],
                "inbound": 0,
                "outbound": 0,
                "adjustment": 0,
                "net": 0,
                "current_stock": log["current_stock"],
            }
        if log["operation_type"] == "inbound":
            summary[key]["inbound"] += abs(parse_int(log["quantity"]))
        elif log["operation_type"] == "outbound":
            summary[key]["outbound"] += abs(parse_int(log["quantity"]))
        else:
            summary[key]["adjustment"] += delta
        summary[key]["net"] += delta

    summary_sheet = workbook.create_sheet("商品別集計")
    summary_sheet.append([
        "品番",
        "大分類",
        "メーカー/商品",
        "概要",
        "入庫数",
        "出庫数",
        "棚卸/修正増減",
        "純増減",
        "現在販売可能台数",
    ])
    for row in summary.values():
        summary_sheet.append([
            row["sku"],
            row["big_category"],
            row["maker_or_product"],
            row["overview"],
            row["inbound"],
            row["outbound"],
            row["adjustment"],
            row["net"],
            row["current_stock"],
        ])

    style_report_sheet(detail_sheet)
    style_report_sheet(summary_sheet)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output, start


@app.route("/")
@login_required
def dashboard():
    conn = get_db_connection()
    today = datetime.today().date()
    cutoff = today - timedelta(days=30)

    zero_stock = conn.execute(
        "SELECT p.* FROM zero_stock_alerts za JOIN products p ON za.product_id = p.id WHERE p.current_stock = 0 ORDER BY p.name"
    ).fetchall()

    low_stock = conn.execute(
        "SELECT p.*, la.threshold FROM low_stock_alerts la JOIN products p ON la.product_id = p.id WHERE p.current_stock > 0 AND p.current_stock <= la.threshold ORDER BY p.current_stock ASC"
    ).fetchall()

    stagnant_stock = conn.execute(
        "SELECT p.* FROM stagnant_stock_alerts sa JOIN products p ON sa.product_id = p.id WHERE p.current_stock > 0 AND p.last_out_date IS NOT NULL AND date(p.last_out_date) < date(?) ORDER BY p.last_out_date"
    , (cutoff.isoformat(),)).fetchall()

    recent_logs = conn.execute(
        "SELECT stock_logs.*, products.name, products.sku FROM stock_logs JOIN products ON stock_logs.product_id = products.id ORDER BY stock_logs.created_at DESC LIMIT 10"
    ).fetchall()

    conn.close()
    return render_template(
        "dashboard.html",
        zero_stock=zero_stock,
        low_stock=low_stock,
        stagnant_stock=stagnant_stock,
        recent_logs=recent_logs,
        cutoff=cutoff,
        operation_labels=OPERATION_LABELS,
    )


@app.route("/alerts/<alert_type>", methods=["GET", "POST"])
@login_required
def alert_settings(alert_type):
    if alert_type not in ["low_stock", "zero_stock", "stagnant_stock"]:
        return "指定されたアラートタイプが無効です。", 404

    params = parse_search_args()
    conn = get_db_connection()

    if request.method == "POST":
        if g.user["role"] != "admin":
            conn.close()
            flash("管理者権限が必要です。", "danger")
            return redirect(url_for("alert_settings", alert_type=alert_type, **params))
        action = request.form.get("action")
        product_id = request.form.get("product_id", type=int)

        if action == "add_low_stock":
            threshold = parse_int(request.form.get("threshold"))
            if product_id is None or threshold < 0:
                flash("有効な商品と最低在庫を指定してください。", "danger")
            else:
                conn.execute(
                    "INSERT OR REPLACE INTO low_stock_alerts (product_id, threshold) VALUES (?, ?)",
                    (product_id, threshold),
                )
                conn.commit()
                flash("低在庫リマインダーを追加しました。", "success")
        elif action == "remove_low_stock":
            if product_id is not None:
                conn.execute("DELETE FROM low_stock_alerts WHERE product_id = ?", (product_id,))
                conn.commit()
                flash("低在庫リマインダーを削除しました。", "success")
        elif action == "add_zero_stock":
            if product_id is None:
                flash("有効な商品を指定してください。", "danger")
            else:
                conn.execute(
                    "INSERT OR IGNORE INTO zero_stock_alerts (product_id) VALUES (?)",
                    (product_id,),
                )
                conn.commit()
                flash("在庫0リマインダーを追加しました。", "success")
        elif action == "remove_zero_stock":
            if product_id is not None:
                conn.execute("DELETE FROM zero_stock_alerts WHERE product_id = ?", (product_id,))
                conn.commit()
                flash("在庫0リマインダーを削除しました。", "success")
        elif action == "add_stagnant_stock":
            if product_id is None:
                flash("有効な商品を指定してください。", "danger")
            else:
                conn.execute(
                    "INSERT OR IGNORE INTO stagnant_stock_alerts (product_id) VALUES (?)",
                    (product_id,),
                )
                conn.commit()
                flash("不動在庫リマインダーを追加しました。", "success")
        elif action == "remove_stagnant_stock":
            if product_id is not None:
                conn.execute("DELETE FROM stagnant_stock_alerts WHERE product_id = ?", (product_id,))
                conn.commit()
                flash("不動在庫リマインダーを削除しました。", "success")
        conn.close()
        return redirect(url_for("alert_settings", alert_type=alert_type, **params))

    if alert_type == "low_stock":
        alert_products = conn.execute(
            "SELECT la.product_id AS alert_product_id, la.threshold, p.* FROM low_stock_alerts la JOIN products p ON la.product_id = p.id ORDER BY p.name"
        ).fetchall()
    elif alert_type == "zero_stock":
        alert_products = conn.execute(
            "SELECT za.product_id AS alert_product_id, p.* FROM zero_stock_alerts za JOIN products p ON za.product_id = p.id ORDER BY p.name"
        ).fetchall()
    else:
        alert_products = conn.execute(
            "SELECT sa.product_id AS alert_product_id, p.* FROM stagnant_stock_alerts sa JOIN products p ON sa.product_id = p.id ORDER BY p.name"
        ).fetchall()

    sql, values = build_product_query(params)
    search_results = conn.execute(sql + " LIMIT 50", values).fetchall()
    conn.close()

    selected_product_ids = {row["alert_product_id"] for row in alert_products}

    return render_template(
        "alert_settings.html",
        alert_type=alert_type,
        alert_products=alert_products,
        search_results=search_results,
        selected_product_ids=selected_product_ids,
        params=params,
    )


@app.route("/stock/log/edit/<int:log_id>", methods=["GET", "POST"])
@admin_required
def edit_stock_log(log_id):
    params = parse_search_args()
    selected_product_id = request.args.get("selected_product_id", type=int)

    conn = get_db_connection()
    log = conn.execute(
        "SELECT stock_logs.*, products.name, products.sku, products.current_stock AS product_current_stock FROM stock_logs JOIN products ON stock_logs.product_id = products.id WHERE stock_logs.id = ?",
        (log_id,),
    ).fetchone()
    if log is None:
        conn.close()
        return "ログが見つかりません。", 404

    old_product = conn.execute(
        "SELECT * FROM products WHERE id = ?",
        (log["product_id"],),
    ).fetchone()

    search_sql, search_values = build_product_query(params)
    search_results = conn.execute(search_sql + " LIMIT 50", search_values).fetchall()

    if selected_product_id:
        selected_product = conn.execute(
            "SELECT * FROM products WHERE id = ?",
            (selected_product_id,),
        ).fetchone()
        if selected_product is None:
            selected_product = old_product
    else:
        selected_product = old_product

    selected_product_id = selected_product["id"]

    error_message = None
    if request.method == "POST":
        action = request.form.get("action", "save")
        if action == "delete_log":
            restored_snapshot = False
            metadata = {}
            if log["metadata_json"]:
                try:
                    metadata = json.loads(log["metadata_json"])
                except Exception:
                    metadata = {}

            if log["operation_type"] == "modification" and metadata.get("kind") == "inventory_edit":
                restored_snapshot = restore_product_snapshot(
                    conn,
                    old_product["id"],
                    metadata.get("before", {}),
                )

            if not restored_snapshot:
                old_delta = compute_stock_delta(log["operation_type"], log["quantity"])
                new_stock = old_product["current_stock"] - old_delta
                staff_stock_json = old_product["staff_stock_json"]
                if log["operation_type"] == "outbound":
                    staff_stock_json = adjust_staff_sales_json(
                        staff_stock_json,
                        log["staff_name"],
                        -abs(log["quantity"]),
                    )
                conn.execute(
                    "UPDATE products SET current_stock = ?, available_stock = ?, staff_stock_json = ?, updated_at = ? WHERE id = ?",
                    (new_stock, new_stock, staff_stock_json, datetime.now().strftime("%Y-%m-%d %H:%M:%S"), old_product["id"]),
                )
            conn.execute("DELETE FROM stock_logs WHERE id = ?", (log_id,))
            conn.commit()
            conn.close()
            if log["operation_type"] == "modification" and not restored_snapshot:
                flash("在庫記録を削除しました。旧形式の修正記録のため、数量だけ回復しました。", "success")
            else:
                flash("在庫記録を削除しました。", "success")
            return redirect(url_for("dashboard"))

        target_product_id = request.form.get("product_id", type=int)
        operation_type = request.form.get("operation_type", log["operation_type"])
        quantity = request.form.get("quantity", type=int)
        staff_name = request.form.get("staff_name", "").strip()
        note = request.form.get("note", "").strip()

        if target_product_id is None:
            error_message = "商品を選択してください。"
        elif operation_type not in ["inbound", "outbound", "adjustment", "modification"]:
            error_message = "有効な操作タイプを選択してください。"
        elif quantity is None:
            error_message = "数量を入力してください。"
        elif quantity < 0 and operation_type in ["inbound", "outbound"]:
            error_message = "入庫または出庫の数量は0以上で入力してください。"
        else:
            old_delta = compute_stock_delta(log["operation_type"], log["quantity"])
            new_delta = compute_stock_delta(operation_type, quantity)
            if operation_type == "outbound":
                quantity = abs(quantity)
                new_delta = compute_stock_delta(operation_type, quantity)
            log_quantity = abs(quantity) if operation_type in ["inbound", "outbound"] else quantity

            if target_product_id == old_product["id"]:
                new_stock = old_product["current_stock"] + (new_delta - old_delta)
                if new_stock < 0:
                    error_message = "在庫数が負になります。数量を調整してください。"
                else:
                    staff_stock_json = old_product["staff_stock_json"]
                    if log["operation_type"] == "outbound":
                        staff_stock_json = adjust_staff_sales_json(
                            staff_stock_json,
                            log["staff_name"],
                            -abs(log["quantity"]),
                        )
                    if operation_type == "outbound":
                        staff_stock_json = adjust_staff_sales_json(
                            staff_stock_json,
                            staff_name,
                            abs(quantity),
                        )
                    conn.execute(
                        "UPDATE products SET current_stock = ?, available_stock = ?, staff_stock_json = ?, updated_at = ? WHERE id = ?",
                        (new_stock, new_stock, staff_stock_json, datetime.now().strftime("%Y-%m-%d %H:%M:%S"), old_product["id"]),
                    )
            else:
                target_product = conn.execute(
                    "SELECT * FROM products WHERE id = ?",
                    (target_product_id,),
                ).fetchone()
                if target_product is None:
                    error_message = "選択した商品が見つかりません。"
                else:
                    new_stock_old = old_product["current_stock"] - old_delta
                    new_stock_target = target_product["current_stock"] + new_delta
                    if new_stock_old < 0 or new_stock_target < 0:
                        error_message = "在庫数が負になります。商品と数量を確認してください。"
                    else:
                        old_staff_stock_json = old_product["staff_stock_json"]
                        if log["operation_type"] == "outbound":
                            old_staff_stock_json = adjust_staff_sales_json(
                                old_staff_stock_json,
                                log["staff_name"],
                                -abs(log["quantity"]),
                            )
                        target_staff_stock_json = target_product["staff_stock_json"]
                        if operation_type == "outbound":
                            target_staff_stock_json = adjust_staff_sales_json(
                                target_staff_stock_json,
                                staff_name,
                                abs(quantity),
                            )
                        conn.execute(
                            "UPDATE products SET current_stock = ?, available_stock = ?, staff_stock_json = ?, updated_at = ? WHERE id = ?",
                            (new_stock_old, new_stock_old, old_staff_stock_json, datetime.now().strftime("%Y-%m-%d %H:%M:%S"), old_product["id"]),
                        )
                        conn.execute(
                            "UPDATE products SET current_stock = ?, available_stock = ?, staff_stock_json = ?, updated_at = ? WHERE id = ?",
                            (new_stock_target, new_stock_target, target_staff_stock_json, datetime.now().strftime("%Y-%m-%d %H:%M:%S"), target_product_id),
                        )

            if error_message is None:
                conn.execute(
                    "UPDATE stock_logs SET product_id = ?, operation_type = ?, quantity = ?, staff_name = ?, note = ? WHERE id = ?",
                    (target_product_id, operation_type, log_quantity, staff_name, note, log_id),
                )
                conn.commit()
                conn.close()
                flash("在庫記録を更新しました。", "success")
                return redirect(url_for("dashboard"))

    conn.close()
    return render_template(
        "edit_stock_log.html",
        log=log,
        operation_label=OPERATION_LABELS.get(log["operation_type"], log["operation_type"]),
        error_message=error_message,
        search_results=search_results,
        params=params,
        selected_product=selected_product,
        selected_product_id=selected_product_id,
    )


@app.route("/inventory")
@login_required
def inventory():
    params = parse_search_args()
    edit_mode = g.user["role"] == "admin" and request.args.get("edit") == "1"
    sql, values = build_product_query(params)
    conn = get_db_connection()
    products = conn.execute(sql, values).fetchall()
    
    # Get all distinct big_categories
    big_categories_result = conn.execute(
        "SELECT DISTINCT big_category FROM products WHERE big_category IS NOT NULL AND big_category != '' ORDER BY big_category"
    ).fetchall()
    big_categories = [row[0] for row in big_categories_result]
    
    conn.close()
    return render_template(
        "inventory.html",
        products=products,
        params=params,
        big_categories=big_categories,
        edit_mode=edit_mode,
    )


@app.route("/inventory/new", methods=["GET", "POST"])
@admin_required
def inventory_new():
    conn = get_db_connection()
    options = get_inventory_form_options(conn)
    error_message = None
    form_data = {}
    staff_rows = staff_stock_form_rows()

    if request.method == "POST":
        form_data = request.form.to_dict()
        big_category = request.form.get("big_category", "").strip()
        maker_or_product = request.form.get("maker_or_product", "").strip()
        overview = request.form.get("overview", "").strip()
        sku = request.form.get("sku", "").strip()
        supplier = request.form.get("supplier", "").strip()
        amount = request.form.get("amount", "").strip()
        available_stock = parse_int(request.form.get("available_stock"))
        total_stock = parse_int(request.form.get("total_stock"))
        notes = request.form.get("notes", "").strip()
        source_sheet = request.form.get("source_sheet", "").strip() or "手動追加"
        staff_stock = parse_staff_stock_form()
        staff_rows = staff_stock_form_rows(json.dumps(staff_stock, ensure_ascii=False))

        if not big_category:
            error_message = "大分類を入力してください。"
        elif not sku:
            error_message = "品番を入力してください。"
        elif available_stock < 0 or total_stock < 0:
            error_message = "台数は0以上で入力してください。"
        else:
            existing = conn.execute("SELECT id FROM products WHERE sku = ?", (sku,)).fetchone()
            if existing:
                error_message = "同じ品番の商品がすでにあります。"

        if error_message is None:
            source_row = conn.execute(
                "SELECT COALESCE(MAX(source_row), 0) + 1 FROM products WHERE source_sheet = ?",
                (source_sheet,),
            ).fetchone()[0]
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            product_name = maker_or_product or sku or overview or supplier or "不明"
            cursor = conn.execute(
                "INSERT INTO products (source_sheet, source_row, big_category, maker_or_product, overview, supplier, amount, sku, stock_status, display_flag, available_stock, total_stock, reorder_point, staff_stock_json, notes, imported_at, current_stock, name, category, location, created_at, updated_at) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                (
                    source_sheet,
                    source_row,
                    big_category,
                    maker_or_product,
                    overview,
                    supplier,
                    amount,
                    sku,
                    "",
                    "",
                    available_stock,
                    total_stock,
                    0,
                    json.dumps(staff_stock, ensure_ascii=False),
                    notes,
                    now,
                    available_stock,
                    product_name,
                    big_category,
                    "",
                    now,
                    now,
                ),
            )
            conn.execute(
                "INSERT INTO stock_logs (product_id, operation_type, quantity, staff_name, note, created_at) VALUES (?, ?, ?, ?, ?, ?)",
                (cursor.lastrowid, "modification", available_stock, "", "新品追加", now),
            )
            conn.commit()
            conn.close()
            flash("新品を追加しました。", "success")
            return redirect(url_for("inventory"))

    conn.close()
    return render_template(
        "inventory_new.html",
        title="新品追加",
        submit_label="追加する",
        options=options,
        error_message=error_message,
        form_data=form_data,
        staff_rows=staff_rows,
    )


@app.route("/inventory/edit/<int:product_id>", methods=["GET", "POST"])
@admin_required
def inventory_edit(product_id):
    conn = get_db_connection()
    product = conn.execute("SELECT * FROM products WHERE id = ?", (product_id,)).fetchone()
    if product is None:
        conn.close()
        return "商品が見つかりません。", 404

    options = get_inventory_form_options(conn)
    error_message = None
    form_data = dict(product)
    staff_rows = staff_stock_form_rows(product["staff_stock_json"])

    if request.method == "POST":
        form_data = request.form.to_dict()
        big_category = request.form.get("big_category", "").strip()
        maker_or_product = request.form.get("maker_or_product", "").strip()
        overview = request.form.get("overview", "").strip()
        supplier = request.form.get("supplier", "").strip()
        amount = request.form.get("amount", "").strip()
        sku = request.form.get("sku", "").strip()
        available_stock = parse_int(request.form.get("available_stock"))
        total_stock = parse_int(request.form.get("total_stock"))
        notes = request.form.get("notes", "").strip()
        source_sheet = request.form.get("source_sheet", "").strip() or "手動追加"
        staff_stock = parse_staff_stock_form()
        staff_rows = staff_stock_form_rows(json.dumps(staff_stock, ensure_ascii=False))

        if not big_category:
            error_message = "大分類を入力してください。"
        elif not sku:
            error_message = "品番を入力してください。"
        elif available_stock < 0 or total_stock < 0:
            error_message = "台数は0以上で入力してください。"
        else:
            existing = conn.execute(
                "SELECT id FROM products WHERE sku = ? AND id != ?",
                (sku, product_id),
            ).fetchone()
            if existing:
                error_message = "同じ品番の商品がすでにあります。"

        if error_message is None:
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            stock_change = available_stock - product["current_stock"]
            product_name = maker_or_product or sku or overview or supplier or "不明"
            before_snapshot = product_snapshot(product)
            after_snapshot = {
                **before_snapshot,
                "big_category": big_category,
                "maker_or_product": maker_or_product,
                "overview": overview,
                "supplier": supplier,
                "amount": amount,
                "sku": sku,
                "available_stock": available_stock,
                "current_stock": available_stock,
                "total_stock": total_stock,
                "staff_stock_json": json.dumps(staff_stock, ensure_ascii=False),
                "notes": notes,
                "source_sheet": source_sheet,
                "name": product_name,
                "category": big_category,
            }
            conn.execute(
                """
                UPDATE products
                SET big_category = ?, maker_or_product = ?, overview = ?, supplier = ?,
                    amount = ?, sku = ?, available_stock = ?, current_stock = ?,
                    total_stock = ?, staff_stock_json = ?, notes = ?, source_sheet = ?,
                    name = ?, category = ?, updated_at = ?
                WHERE id = ?
                """,
                (
                    big_category,
                    maker_or_product,
                    overview,
                    supplier,
                    amount,
                    sku,
                    available_stock,
                    available_stock,
                    total_stock,
                    json.dumps(staff_stock, ensure_ascii=False),
                    notes,
                    source_sheet,
                    product_name,
                    big_category,
                    now,
                    product_id,
                ),
            )
            conn.execute(
                "INSERT INTO stock_logs (product_id, operation_type, quantity, staff_name, note, created_at, metadata_json) VALUES (?, ?, ?, ?, ?, ?, ?)",
                (
                    product_id,
                    "modification",
                    stock_change,
                    "",
                    "在庫修正",
                    now,
                    json.dumps(
                        {
                            "kind": "inventory_edit",
                            "before": before_snapshot,
                            "after": after_snapshot,
                        },
                        ensure_ascii=False,
                    ),
                ),
            )
            conn.commit()
            conn.close()
            flash("在庫情報を更新しました。", "success")
            return redirect(url_for("inventory", edit="1"))

    conn.close()
    return render_template(
        "inventory_new.html",
        title="在庫修正",
        submit_label="更新する",
        options=options,
        error_message=error_message,
        form_data=form_data,
        staff_rows=staff_rows,
        is_edit=True,
    )


@app.route("/stocktaking")
@login_required
def stocktaking():
    params = parse_search_args()
    sql, values = build_product_query(params)
    conn = get_db_connection()
    products = conn.execute(sql, values).fetchall()
    big_categories_result = conn.execute(
        "SELECT DISTINCT big_category FROM products WHERE big_category IS NOT NULL AND big_category != '' ORDER BY big_category"
    ).fetchall()
    big_categories = [row[0] for row in big_categories_result]
    conn.close()
    return render_template(
        "stocktaking.html",
        products=products,
        params=params,
        big_categories=big_categories,
        is_started=session.get("stocktaking_started", False),
        checked_product_ids=set(session.get("stocktaking_checked_ids", [])),
    )


@app.route("/stocktaking/start", methods=["POST"])
@admin_required
def stocktaking_start():
    session["stocktaking_started"] = True
    session["stocktaking_checked_ids"] = []
    flash("棚卸を開始しました。", "success")
    return redirect(url_for("stocktaking", **parse_search_args()))


@app.route("/stocktaking/end", methods=["POST"])
@admin_required
def stocktaking_end():
    session["stocktaking_started"] = False
    session["stocktaking_checked_ids"] = []
    flash("棚卸を終了しました。", "success")
    return redirect(url_for("stocktaking", **parse_search_args()))


@app.route("/stocktaking/check/<int:product_id>", methods=["POST"])
@admin_required
def stocktaking_check(product_id):
    conn = get_db_connection()
    product = conn.execute("SELECT id FROM products WHERE id = ?", (product_id,)).fetchone()
    conn.close()
    if product is None:
        return "商品が見つかりません。", 404
    checked_ids = set(session.get("stocktaking_checked_ids", []))
    checked_ids.add(product_id)
    session["stocktaking_checked_ids"] = list(checked_ids)
    flash("棚卸チェックを確認しました。", "success")
    return redirect(url_for("stocktaking", **parse_search_args()) + f"#product-{product_id}")


@app.route("/stocktaking/edit/<int:product_id>", methods=["GET", "POST"])
@admin_required
def stocktaking_edit(product_id):
    conn = get_db_connection()
    product = conn.execute("SELECT * FROM products WHERE id = ?", (product_id,)).fetchone()
    if product is None:
        conn.close()
        return "商品が見つかりません。", 404

    error_message = None
    if request.method == "POST":
        new_stock = request.form.get("new_stock", type=int)
        reason = request.form.get("reason", "").strip()
        if new_stock is None or new_stock < 0:
            error_message = "正しい在庫数を入力してください。"
        elif not reason:
            error_message = "理由を入力してください。"
        else:
            change = new_stock - product["current_stock"]
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            conn.execute(
                "UPDATE products SET current_stock = ?, available_stock = ?, updated_at = ? WHERE id = ?",
                (new_stock, new_stock, now, product_id),
            )
            conn.execute(
                "INSERT INTO stock_logs (product_id, operation_type, quantity, staff_name, note, created_at) VALUES (?, ?, ?, ?, ?, ?)",
                (product_id, "adjustment", change, "", reason, now),
            )
            conn.commit()
            conn.close()
            flash("棚卸の修正を記録しました。", "success")
            return redirect(url_for("stocktaking", **parse_search_args()))

    conn.close()
    return render_template(
        "stocktaking_edit.html",
        product=product,
        error_message=error_message,
    )


@app.route("/excel_import", methods=["GET", "POST"])
@admin_required
def excel_import():
    result = None
    if request.method == "POST":
        uploaded_file = request.files.get("excel_file")
        if not uploaded_file or uploaded_file.filename == "":
            flash("Excelファイルを選択してください。", "danger")
        elif not uploaded_file.filename.lower().endswith(".xlsx"):
            flash(".xlsxファイルをアップロードしてください。", "danger")
        else:
            result = import_excel_file(uploaded_file)
            deleted_msg = f"旧データ：{result['deleted_products']}件の商品と{result['deleted_logs']}件のログを削除しました。"
            if result["errors"] == 0:
                flash(f"Excel取込が完了しました。{deleted_msg} 新規登録：{result['imported']}件、スキップ：{result['skipped']}件。", "success")
            else:
                flash(f"Excel取込が完了しました。{deleted_msg} 新規登録：{result['imported']}件、スキップ：{result['skipped']}件、エラー：{result['errors']}件。", "danger")
    return render_template("excel_import.html", result=result)


@app.route("/excel_export")
@admin_required
def excel_export():
    return render_template("excel_export.html")


@app.route("/monthly_changes_export")
@admin_required
def monthly_changes_export():
    output, start = export_monthly_changes_file()
    filename = f"monthly_changes_{start.strftime('%Y_%m')}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/stock/operate/<int:product_id>", methods=["GET", "POST"])
@admin_required
def stock_operate(product_id):
    operation_type = request.args.get("type", "inbound")
    if operation_type not in ["inbound", "outbound", "adjustment", "modification"]:
        operation_type = "inbound"

    conn = get_db_connection()
    product = conn.execute(
        "SELECT * FROM products WHERE id = ?", (product_id,)
    ).fetchone()

    if product is None:
        conn.close()
        return "商品が見つかりません。", 404

    error_message = None
    if request.method == "POST":
        staff_name = request.form.get("staff_name", "").strip()
        note = request.form.get("note", "").strip()
        today = datetime.today().date().isoformat()
        current_stock = product["current_stock"]
        quantity = None
        change = None
        last_in_date = product["last_in_date"]
        last_out_date = product["last_out_date"]

        if operation_type == "adjustment":
            actual_quantity = request.form.get("actual_quantity", type=int)
            if actual_quantity is None or actual_quantity < 0:
                error_message = "正しい実際在庫数を入力してください。"
            else:
                change = actual_quantity - current_stock
                quantity = change
                new_stock = actual_quantity
        elif operation_type == "modification":
            new_stock = request.form.get("new_stock", type=int)
            if new_stock is None or new_stock < 0:
                error_message = "正しい新在庫数を入力してください。"
            else:
                change = new_stock - current_stock
                quantity = change
        else:
            quantity = request.form.get("quantity", type=int)
            if quantity is None or quantity < 0:
                error_message = "正しい数量を入力してください。"
            else:
                if operation_type == "inbound":
                    change = quantity
                    new_stock = current_stock + quantity
                    last_in_date = today
                elif operation_type == "outbound":
                    if quantity > current_stock:
                        error_message = "在庫が不足しています。"
                    else:
                        change = -quantity
                        new_stock = current_stock - quantity
                        last_out_date = today

        if error_message is None and change is not None:
            staff_stock_json = product["staff_stock_json"]
            if operation_type == "outbound":
                staff_stock_json = adjust_staff_sales_json(staff_stock_json, staff_name, abs(quantity))
            conn.execute(
                "UPDATE products SET current_stock = ?, available_stock = ?, staff_stock_json = ?, last_in_date = ?, last_out_date = ?, updated_at = ? WHERE id = ?",
                (new_stock, new_stock, staff_stock_json, last_in_date, last_out_date, today + " 00:00:00", product_id),
            )
            conn.execute(
                "INSERT INTO stock_logs (product_id, operation_type, quantity, staff_name, note, created_at) VALUES (?, ?, ?, ?, ?, ?)",
                (product_id, operation_type, abs(quantity) if operation_type in ["inbound", "outbound"] else quantity, staff_name, note, datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            )
            conn.commit()
            conn.close()
            flash("在庫操作が正常に完了しました。", "success")
            return redirect(url_for("inventory"))

    conn.close()
    return render_template(
        "stock_operation.html",
        product=product,
        operation_type=operation_type,
        operation_label=OPERATION_LABELS.get(operation_type, operation_type),
        error_message=error_message,
    )


if __name__ == "__main__":
    host = os.getenv("APP_HOST", "127.0.0.1")
    port = int(os.getenv("APP_PORT", "5001"))
    app.run(debug=True, host=host, port=port)
